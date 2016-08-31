import openpyxl as x
import openpyxl.styles as xstyles
from math import sqrt

fz_path = '/Users/chrisprobst/Desktop/Fz.xlsx'
f4_path = '/Users/chrisprobst/Desktop/F4.xlsx'
f3_path = '/Users/chrisprobst/Desktop/F3.xlsx'

b2_f3_path = '/Users/chrisprobst/Desktop/B2_F3.xlsx'
b2_f4_path = '/Users/chrisprobst/Desktop/B2_F4.xlsx'
b2_fz_path = '/Users/chrisprobst/Desktop/B2_Fz.xlsx'

input_path = '/Users/chrisprobst/Desktop/EEG_Auswertung.xlsx'
output_path = '/Users/chrisprobst/Desktop/EEG_Auswertung_generated.xlsx'

################################################################################
################################################################################

fzWB = x.load_workbook(fz_path)
f4WB = x.load_workbook(f4_path)
f3WB = x.load_workbook(f3_path)

b2f3WB = x.load_workbook(b2_f3_path)
b2f4WB = x.load_workbook(b2_f4_path)
b2fzWB = x.load_workbook(b2_fz_path)

inputWB = x.load_workbook(input_path)

################################################################################
################################################################################

redFill = xstyles.PatternFill(start_color='FFFF0000',
                              end_color='FFFF0000',
                              fill_type='solid')

orangeFill = xstyles.PatternFill(start_color='FFFFC000',
                                 end_color='FFFFC000',
                                 fill_type='solid')

greenFill = xstyles.PatternFill(start_color='FF00FF00',
                                end_color='FF00FF00',
                                fill_type='solid')
# Beta1+3
row_alpha_offset = 6
row_beta1_offset = 7
row_beta3_offset = 8

# Beta2
row_beta2_offset = 6
row_theta_offset = 7
row_delta_offset = 8

row_percentage_offset = 10
row_step = 12

################################################################################
############################### Build the cache ################################
################################################################################

row_index_cache = {}
row_sub_index_cache = {}
row_final_index_cache = {}
column_index_cache = {}
result_rows = inputWB.active.rows[:]
column_names = result_rows[0][2:]
data_rows = result_rows[1:]

def build_caches():
    for i, row in enumerate(data_rows[:43]):
        row_index_cache[row[0].value] = i+1
        row_sub_index_cache[row[0].value] = i+1+43
        row_final_index_cache[row[0].value] = i+1+43+43

    for i, column in enumerate(column_names):
        column_index_cache[column.value] = i+2

# Build caches
build_caches()

################################################################################
############################### Insert value into output #######################
################################################################################

def insert_value_into_output(row_index, column_index, percentage, new_value):
    if row_index not in row_index_cache:
        print("[VP%d] Could not find row: %d" % (row_index, row_index))
        return

    if column_index not in column_index_cache:
        print("[VP%d] Could not find column: %s" % (row_index, column_index))
        return

    i = row_index_cache[row_index]
    j = column_index_cache[column_index]
    c = result_rows[i][j]
    c.value = new_value

    if percentage > 100:
        c.fill = redFill
    elif percentage > 40:
        c.fill = orangeFill
    elif percentage > 30:
        c.fill = greenFill

def insert_sub_value_into_output(row_index, column_index, percentage, new_value):
    if row_index not in row_sub_index_cache:
        print("[VP%d] Could not find row: %d" % (row_index, row_index))
        return

    if column_index not in column_index_cache:
        print("[VP%d] Could not find column: %s" % (row_index, column_index))
        return

    i = row_sub_index_cache[row_index]
    j = column_index_cache[column_index]
    c = result_rows[i][j]
    c.value = new_value

    if percentage > 100:
        c.fill = redFill
    elif percentage > 40:
        c.fill = orangeFill
    elif percentage > 30:
        c.fill = greenFill

def insert_final_value_into_output(row_index, column_index, percentage, new_value):
    if row_index not in row_final_index_cache:
        print("[VP%d] Could not find row: %d" % (row_index, row_index))
        return

    if column_index not in column_index_cache:
        print("[VP%d] Could not find column: %s" % (row_index, column_index))
        return

    i = row_final_index_cache[row_index]
    j = column_index_cache[column_index]
    c = result_rows[i][j]
    c.value = new_value

    if percentage > 100:
        c.fill = redFill
    elif percentage > 40:
        c.fill = orangeFill
    elif percentage > 30:
        c.fill = greenFill

################################################################################
############################### Copy from F to output ##########################
################################################################################

def copy_from_f_to_output_beta_2(category, wb):
    print('Processing %s...' % category)

    for worksheet in wb.worksheets:
        row_index = int(worksheet.title[2:4])
        row_offset = 5
        rows = worksheet.rows[:]

        base1_b2 = None
        base2_b2 = None
        base3_b2 = None

        mean_b2 = None
        var_v2 = None
        std_v2 = None

        counter = 0

        is_special = False
        while True:
            if len(rows) < row_offset:
                break

            table_name = rows[row_offset][0]
            table_beta2_mean = rows[row_offset+row_beta2_offset][2]
            table_theta_mean = rows[row_offset+row_theta_offset][2]
            table_delta_mean = rows[row_offset+row_delta_offset][2]
            table_percentage = rows[row_offset+row_percentage_offset][0]

            table_name_value = table_name.value

            if table_name_value == None:
                row_offset += row_step
                continue

            sub_result = table_name_value.split(':')[1]
            if sub_result == None:
                row_offset += row_step
                continue

            table_name_value = sub_result.split('(')[0][:-1]

            if 'Baseline' not in table_name_value and 'Base' in table_name_value:
                table_name_value = table_name_value.replace('Base', 'B')
                is_special = True
            elif 'Baseline' in table_name_value:
                table_name_value = table_name_value.replace('Baseline', 'B')

            if is_special:
                table_name_value = table_name_value.replace('_', '')
            else:
                table_name_value = table_name_value.replace('.', '')

            if table_name_value == 'B1':
                base1_b2 = float(table_beta2_mean.value.strip())
                counter += 1
            elif table_name_value == 'B2':
                base2_b2 = float(table_beta2_mean.value.strip())
                counter += 1
            elif table_name_value == 'B3':
                base3_b2 = float(table_beta2_mean.value.strip())
                counter += 1

            if std_v2 == None and counter == 3:
                mean_b2 = (base1_b2 + base2_b2 + base3_b2) / 3
                var_v2 = ((base1_b2 - mean_b2) ** 2 + (base2_b2 - mean_b2) ** 2 + (base3_b2 - mean_b2) ** 2) / 3
                std_v2 = sqrt(var_v2) * 3

                insert_value_into_output(row_index, category + 'MB_MEAN', 0, mean_b2)
                insert_value_into_output(row_index, category + 'MB_STD', 0, std_v2)

            if 'Cue' not in table_name_value:
                table_beta2_mean_value = table_beta2_mean.value
                table_theta_mean_value = table_theta_mean.value
                table_delta_mean_value = table_delta_mean.value
                table_percentage_value = table_percentage.value

                percentage = float(table_percentage_value.split('%')[0].strip())

                # Sub-Matrix
                lb_column_index = category + 'LB_' + table_name_value
                lb_cell_value = float(table_theta_mean_value.strip())
                hb_column_index = category + 'HB_' + table_name_value
                hb_cell_value = float(table_delta_mean_value.strip())

                mb_column_index = category + 'MB_' + table_name_value
                mb_cell_value = float(table_beta2_mean_value.strip())

                if counter == 3:
                    if abs(mean_b2 - mb_cell_value) > std_v2:
                        percentage = 1000
                        # print(mean_b2, mb_cell_value, std_v2)
                        # row_offset += row_step
                        # continue

                insert_value_into_output(row_index, mb_column_index, percentage, mb_cell_value)

                # Sub-Matrix
                insert_sub_value_into_output(row_index, lb_column_index, percentage, lb_cell_value)
                insert_sub_value_into_output(row_index, hb_column_index, percentage, hb_cell_value)

            row_offset += row_step

def copy_from_f_to_output(category, wb):
    print('Processing %s...' % category)

    for worksheet in wb.worksheets:
        row_index = int(worksheet.title[2:4])
        row_offset = 5
        rows = worksheet.rows[:]

        base1_b1 = None
        base2_b1 = None
        base3_b1 = None

        base1_b3 = None
        base2_b3 = None
        base3_b3 = None

        mean_b1 = None
        mean_b3 = None

        var_v1 = None
        var_v3 = None

        std_v1 = None
        std_v3 = None

        counter = 0

        is_special = False
        while True:
            if len(rows) < row_offset:
                break

            table_name = rows[row_offset][0]
            table_alpha_mean = rows[row_offset+row_alpha_offset][2]
            table_beta1_mean = rows[row_offset+row_beta1_offset][2]
            table_beta3_mean = rows[row_offset+row_beta3_offset][2]
            table_percentage = rows[row_offset+row_percentage_offset][0]

            table_name_value = table_name.value
            table_name_value = table_name_value.split(':')[1].split('(')[0][:-1]

            if 'Baseline' not in table_name_value and 'Base' in table_name_value:
                table_name_value = table_name_value.replace('Base', 'B')
                is_special = True
            elif 'Baseline' in table_name_value:
                table_name_value = table_name_value.replace('Baseline', 'B')

            if is_special:
                table_name_value = table_name_value.replace('_', '')
            else:
                table_name_value = table_name_value.replace('.', '')

            if table_name_value == 'B1':
                base1_b1 = float(table_beta1_mean.value.strip())
                base1_b3 = float(table_beta3_mean.value.strip())
                counter += 1
            elif table_name_value == 'B2':
                base2_b1 = float(table_beta1_mean.value.strip())
                base2_b3 = float(table_beta3_mean.value.strip())
                counter += 1
            elif table_name_value == 'B3':
                base3_b1 = float(table_beta1_mean.value.strip())
                base3_b3 = float(table_beta3_mean.value.strip())
                counter += 1

            if std_v1 == None and counter == 3:
                mean_b1 = (base1_b1 + base2_b1 + base3_b1) / 3
                mean_b3 = (base1_b3 + base2_b3 + base3_b3) / 3

                var_v1 = ((base1_b1 - mean_b1) ** 2 + (base2_b1 - mean_b1) ** 2 + (base3_b1 - mean_b1) ** 2) / 3
                var_v3 = ((base1_b3 - mean_b3) ** 2 + (base2_b3 - mean_b3) ** 2 + (base3_b3 - mean_b3) ** 2) / 3

                std_v1 = sqrt(var_v1) * 3
                std_v3 = sqrt(var_v3) * 3

                insert_value_into_output(row_index, category + 'LB_MEAN', 0, mean_b1)
                insert_value_into_output(row_index, category + 'HB_MEAN', 0, mean_b3)

                insert_value_into_output(row_index, category + 'LB_STD', 0, std_v1)
                insert_value_into_output(row_index, category + 'HB_STD', 0, std_v3)

            if 'Cue' not in table_name_value:
                table_alpha_mean_value = table_alpha_mean.value
                table_beta1_mean_value = table_beta1_mean.value
                table_beta3_mean_value = table_beta3_mean.value
                table_percentage_value = table_percentage.value

                percentage = float(table_percentage_value.split('%')[0].strip())

                lb_column_index = category + 'LB_' + table_name_value
                lb_cell_value = float(table_beta1_mean_value.strip())

                hb_column_index = category + 'HB_' + table_name_value
                hb_cell_value = float(table_beta3_mean_value.strip())

                p1 = percentage
                p2 = percentage
                if counter == 3:
                    if abs(mean_b1 - lb_cell_value) > std_v1:
                        p1 = 1000
                        # row_offset += row_step
                        # continue

                    if abs(mean_b3 - hb_cell_value) > std_v3:
                        p2 = 1000
                        # row_offset += row_step
                        # continue

                # Sub-Matrix
                mb_column_index = category + 'MB_' + table_name_value
                mb_cell_value = float(table_alpha_mean_value.strip())

                insert_value_into_output(row_index, lb_column_index, percentage, lb_cell_value)
                insert_value_into_output(row_index, hb_column_index, percentage, hb_cell_value)

                # Sub-Matrix
                insert_sub_value_into_output(row_index, mb_column_index, percentage, mb_cell_value)

            row_offset += row_step

################################################################################
############################### Compute relative power #########################
################################################################################

def compute_relative_power():
    print('Computing relative power...')

    # Delta -> HB
    # Theta -> LB
    # Alpha -> MB

    for i in range(1, 43):
        for j in column_names:
            c = j.value
            r = None
            if 'LB' in c:
                r = 'LB'
            elif 'HB' in c:
                r = 'HB'
            elif 'MB' in c:
                r = 'MB'

            delta = result_rows[row_sub_index_cache[i]] [column_index_cache[c.replace(r, 'HB')]].value
            theta = result_rows[row_sub_index_cache[i]] [column_index_cache[c.replace(r, 'LB')]].value
            alpha = result_rows[row_sub_index_cache[i]] [column_index_cache[c.replace(r, 'MB')]].value

            beta1 = result_rows[row_index_cache[i]] [column_index_cache[c.replace(r, 'LB')]].value
            beta2 = result_rows[row_index_cache[i]] [column_index_cache[c.replace(r, 'MB')]].value
            beta3 = result_rows[row_index_cache[i]] [column_index_cache[c.replace(r, 'HB')]].value

            v = result_rows[row_index_cache[i]] [column_index_cache[c]].value

            try:
                result = v / (delta + theta + alpha + beta1 + beta2 + beta3)
                result_rows[row_final_index_cache[i]] [column_index_cache[c]].value = result
            except:
                pass



    #
    # for worksheet in wb.worksheets:
    #     row_index = int(worksheet.title[2:4])
    #     row_offset = 5
    #     rows = worksheet.rows[:]
    #
    #     is_special = False
    #     while True:
    #         if len(rows) < row_offset:
    #             break
    #
    #         table_name = rows[row_offset][0]
    #         table_beta2_mean = rows[row_offset+row_beta2_offset][2]
    #         table_theta_mean = rows[row_offset+row_theta_offset][2]
    #         table_delta_mean = rows[row_offset+row_delta_offset][2]
    #         table_percentage = rows[row_offset+row_percentage_offset][0]
    #
    #         table_name_value = table_name.value
    #
    #         if table_name_value == None:
    #             row_offset += row_step
    #             continue
    #
    #         sub_result = table_name_value.split(':')[1]
    #         if sub_result == None:
    #             row_offset += row_step
    #             continue
    #
    #         table_name_value = sub_result.split('(')[0][:-1]
    #
    #         if 'Baseline' not in table_name_value and 'Base' in table_name_value:
    #             table_name_value = table_name_value.replace('Base', 'B')
    #             is_special = True
    #         elif 'Baseline' in table_name_value:
    #             table_name_value = table_name_value.replace('Baseline', 'B')
    #
    #         if is_special:
    #             table_name_value = table_name_value.replace('_', '')
    #         else:
    #             table_name_value = table_name_value.replace('.', '')
    #
    #         if 'Cue' not in table_name_value:
    #             table_beta2_mean_value = table_beta2_mean.value
    #             table_theta_mean_value = table_theta_mean.value
    #             table_delta_mean_value = table_delta_mean.value
    #             table_percentage_value = table_percentage.value
    #
    #             percentage = float(table_percentage_value.split('%')[0].strip())
    #
    #             # Sub-Matrix
    #             lb_column_index = category + 'LB_' + table_name_value
    #             lb_cell_value = float(table_theta_mean_value.strip())
    #             hb_column_index = category + 'HB_' + table_name_value
    #             hb_cell_value = float(table_delta_mean_value.strip())
    #
    #             mb_column_index = category + 'MB_' + table_name_value
    #             mb_cell_value = float(table_beta2_mean_value.strip())
    #
    #             insert_value_into_output(row_index, mb_column_index, percentage, mb_cell_value)
    #
    #             # Sub-Matrix
    #             insert_sub_value_into_output(row_index, lb_column_index, percentage, lb_cell_value)
    #             insert_sub_value_into_output(row_index, hb_column_index, percentage, hb_cell_value)
    #
    #         row_offset += row_step

################################################################################
############################### Main app #######################################
################################################################################

copy_from_f_to_output("F3", f3WB)
copy_from_f_to_output("F4", f4WB)
copy_from_f_to_output("FZ", fzWB)

################################################################################
################################################################################

copy_from_f_to_output_beta_2("F3", b2f3WB)
copy_from_f_to_output_beta_2("F4", b2f4WB)
copy_from_f_to_output_beta_2("FZ", b2fzWB)

################################################################################
################################################################################

compute_relative_power()

################################################################################
################################################################################

inputWB.save(output_path)
